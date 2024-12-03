import datetime
import random

from openpyxl import Workbook

names = [
    'Александра',
    'Екатерина',
    'Ирина',
    'Ольга',
    'Дарья',
    'Анна',
    'Светлана',
    'Наталья',
    'Елена',
    'Мария',
    'Татьяна',
    'Юлия',
    'Ангелина',
    'София',
    'Алиса',
    'Виктория',
    'Евгения',
    'Анастасия',
    'Людмила',
    'Валентина',
]

lastnames = [
    'Иванова'
    'Петрова',
    'Смирнова',
    'Кузнецова',
    'Соколова',
    'Попова',
    'Лебедева',
    'Козлова',
    'Игнатова',
    'Морозова',
    'Волкова',
    'Зайцева',
    'Дмитриева',
    'Орлова',
    'Козлова',
    'Белова',
    'Федорова',
    'Семенова',
    'Голубева',
    'Тарасова',
]

def create_excel(quiz):
    wb = Workbook()
    ws = wb.active
    counter = 4
    ws['A1'] = f"Квиз №{quiz['id']}"
    ws['B1'] = quiz['title']

    ws['A3'] = "Имя ученика"
    ws['B3'] = "Балл"
    ws['C3'] = "Время"
    ws['D3'] = "Почта"
    ws['E3'] = "Идентификатор"

    for r in quiz['result']:
        ws[f'A{counter}'] = r['name']
        ws[f'B{counter}'] = f'{r["score"]} / {len(quiz["test"])}'
        ws[f'C{counter}'] = r['total']
        ws[f'D{counter}'] = r['mail']
        ws[f'E{counter}'] = r['id']
        counter += 1

    ws[f'A{counter + 1}'] = f'Дата формирования отчета: {datetime.datetime.now().strftime("%d.%m.20%y")}'
    ws[f'A{counter + 2}'] = f"Сотрудник отдела качества образования - {excel_names_random()}"

    wb.save("static/result.xlsx")


def excel_names_random():
    name = random.choice(names)
    last_name = random.choice(lastnames)
    return f'{name} {last_name}'